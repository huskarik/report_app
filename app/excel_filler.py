import openpyxl
from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import PatternFill, Font, Border


def clear_borders(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        ws.cell(row=row, column=col).border = Border()

def copy_row_style(ws, source_row, target_row, max_col=9):
    for col in range(1, max_col + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)

        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


def insert_block(ws, start_row, data_length, template_row, max_col):
    """
    Универсальная вставка блока.
    Возвращает количество реально вставленных строк.
    """
    if data_length <= 1:
        return 0

    rows_to_insert = data_length - 1
    ws.insert_rows(start_row + 1, rows_to_insert)

    for i in range(1, data_length):
        copy_row_style(ws, template_row, start_row + i, max_col)

    return rows_to_insert


def fill_excel_report(template_path, output_path, sections, project_name, from_date, to_date):
    wb = load_workbook(template_path)
    ws = wb['Товары WB на реализации']

    # --- Шапка ---
    ws.cell(row=1, column=1, value=project_name)
    ws.cell(row=1, column=3, value=f"{from_date} - {to_date}")

    total_shift = 0  # сколько строк уже вставлено выше

    # =====================================================
    # СЕКЦИИ A и B (идут параллельно)
    # =====================================================

    data_a = sections['A']['data']
    data_b = sections['B']['data']
    start_ab = sections['A']['start_row']

    max_len_ab = max(len(data_a), len(data_b))
    inserted_ab = insert_block(ws, start_ab, max_len_ab, start_ab, max_col=9)

    # Заполняем
    for i in range(max_len_ab):
        row = start_ab + i

        # -------- A (1-4) --------
        if i < len(data_a):
            item = data_a[i]
            ws.cell(row=row, column=1, value=item.get('art', ''))
            ws.cell(row=row, column=2, value=item.get('name', ''))
            ws.cell(row=row, column=3, value=item.get('quantity', 0))
            ws.cell(row=row, column=4, value=item.get('price', 0))
        else:
            clear_borders(ws, row, 1, 4)

        # -------- B (5-9) --------
        if i < len(data_b):
            item = data_b[i]
            ws.cell(row=row, column=5, value=item.get('art', ''))
            ws.cell(row=row, column=6, value=item.get('name', ''))
            ws.cell(row=row, column=7, value=item.get('quantity', 0))
            ws.cell(row=row, column=8, value=item.get('price', 0))

            if item.get('NSP'):
                cell = ws.cell(row=row, column=9, value=item.get('NSP'))
                cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                cell.font = Font(size=8, bold=True)
        else:
            clear_borders(ws, row, 5, 9)

    total_shift += inserted_ab

    # =====================================================
    # СЕКЦИЯ C
    # =====================================================

    data_c = sections['C']['data']
    start_c = sections['C']['start_row'] + total_shift

    inserted_c = insert_block(ws, start_c, len(data_c), start_c, max_col=3)

    for i in range(len(data_c)):
        ws.cell(row=start_c + i, column=2, value=data_c[i])

    total_shift += inserted_c

    # =====================================================
    # СЕКЦИЯ D
    # =====================================================

    data_d = sections['D']['data']
    start_d = sections['D']['start_row'] + total_shift

    inserted_d = insert_block(ws, start_d, len(data_d), start_d, max_col=3)

    for i in range(len(data_d)):
        ws.cell(row=start_d + i, column=2, value=data_d[i])

    wb.save(output_path)


# =====================================================
# ТЕСТОВЫЕ ДАННЫЕ
# =====================================================
"""
refounds = [
    {'art': "123", 'name': 'Крем косметический СС-31 50 мл', 'price': -1377.5, 'quantity': -1.0, 'NSP': 'НСП'},
    {'art': "321", 'name': 'Крем косметический Эпобис 50 мл', 'price': -1377.5, 'quantity': -1.0, 'NSP': 'НСП'},
    {'art': "6546RR", 'name': 'Иммуноглобулин человека', 'price': -6550.0, 'quantity': -1.0, 'NSP': 'НСП'}
]

demands = [
    {'art': "АБА", 'name': 'Тест 1', 'price': 10, 'quantity': 1.0},
    {'art': "БИБА", 'name': 'Тест 2', 'price': 20, 'quantity': 2.0},
    {'art': "БОБА", 'name': 'Тест 3', 'price': 30, 'quantity': 3.0},
    {'art': "ВЕРА", 'name': 'Тест 4', 'price': 40, 'quantity': 4.0},
    {'art': "РЕВА", 'name': 'Тест 5', 'price': 50, 'quantity': 5.0},
]

comission_numbers = [
    {'art': "BOBA", 'name': 'TEST234', 'price': -1377.5, 'quantity': -1.0},
]

demands_numbers = [
    "Отгрузка 1", "Отгрузка 2", "Отгрузка 3"
]

returns_numbers = [
    "Отчёт комиссионера № 00001",
    "Возврат покупателя № 000032",
    "Отчёт комиссионера № 00002",
    "Возврат покупателя № 000033",
]

sections = {
    'A': {'start_row': 5, 'data': demands},
    'B': {'start_row': 5, 'data': refounds + comission_numbers},
    'C': {'start_row': 8, 'data': demands_numbers},
    'D': {'start_row': 10, 'data': returns_numbers}
}

fill_excel_report(
    'шаблон.xlsx',
    'отчёт.xlsx',
    sections,
    "ozon",
    "02-04-2026 00:00",
    "04-06-2028 00:00"
)
"""